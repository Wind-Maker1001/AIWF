from __future__ import annotations

import os
import re
from dataclasses import dataclass, replace
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, List, Mapping, Optional, Tuple

from aiwf.registry_domains import normalize_registry_domain, summarize_registry_domains
from aiwf.registry_events import record_registry_event
from aiwf.registry_policy import default_conflict_policy, normalize_conflict_policy
from aiwf.runtime_state import get_runtime_state
from aiwf.registry_utils import infer_caller_module


InputReaderFn = Callable[[str, Dict[str, Any]], Tuple[List[Dict[str, Any]], Dict[str, Any]]]


@dataclass(frozen=True)
class InputReaderRegistration:
    input_format: str
    extensions: Tuple[str, ...]
    loader: InputReaderFn
    domain: Optional[str]
    domain_metadata: Mapping[str, Any]
    source_module: str


def _ext(path: str) -> str:
    return Path(path).suffix.lower()


def _normalize_extension(value: str) -> str:
    ext = str(value or "").strip().lower()
    if not ext:
        return ""
    if ext.startswith("."):
        return ext
    if "/" not in ext and "\\" not in ext and "." not in ext:
        return f".{ext}"
    return _ext(ext)


def _normalize_input_format(name: str) -> str:
    normalized = str(name or "").strip().lower()
    if not normalized:
        raise ValueError("input format must be non-empty")
    return normalized


def _extension_conflicts(
    normalized_format: str,
    normalized_extensions: Tuple[str, ...],
) -> Dict[str, Tuple[InputReaderRegistration, set[str]]]:
    state = get_runtime_state()
    conflicts: Dict[str, Tuple[InputReaderRegistration, set[str]]] = {}
    for ext in normalized_extensions:
        current = state.input_readers_by_extension.get(ext)
        if current is None or current.input_format == normalized_format:
            continue
        registration, captured = conflicts.get(current.input_format, (current, set()))
        captured.add(ext)
        conflicts[current.input_format] = (registration, captured)
    return conflicts


def _ensure_builtin_input_readers() -> None:
    state = get_runtime_state()
    if state.builtins_inputs_registered or state.input_bootstrap_in_progress:
        return
    state.input_bootstrap_in_progress = True
    try:
        from aiwf.domains.ingest import register_builtin_input_domains

        register_builtin_input_domains(_register_input_reader)
        state.builtins_inputs_registered = True
    finally:
        state.input_bootstrap_in_progress = False


def _register_input_reader(
    input_format: str,
    extensions: Iterable[str],
    loader: InputReaderFn,
    *,
    domain: Optional[str] = None,
    domain_metadata: Optional[Mapping[str, Any]] = None,
    source_module: Optional[str] = None,
    on_conflict: Optional[str] = None,
) -> InputReaderRegistration:
    state = get_runtime_state()
    normalized_format = _normalize_input_format(input_format)
    normalized_extensions = tuple(
        sorted({ext for ext in (_normalize_extension(value) for value in extensions) if ext})
    )
    if not normalized_extensions:
        raise ValueError("register_input_reader requires at least one extension")
    if not callable(loader):
        raise TypeError("input reader loader must be callable")
    normalized_domain, normalized_domain_metadata = normalize_registry_domain(
        domain,
        domain_metadata,
    )
    source = str(source_module or infer_caller_module())
    policy = normalize_conflict_policy(on_conflict, default_conflict_policy())
    existing = state.input_readers_by_format.get(normalized_format)
    if existing is not None:
        if policy == "error":
            record_registry_event(
                registry="input_reader",
                name=normalized_format,
                action="error",
                policy=policy,
                existing_source_module=existing.source_module,
                new_source_module=source,
                detail="registration already exists",
            )
            raise RuntimeError(
                f"input reader {normalized_format} already registered by {existing.source_module}"
            )
        if policy == "keep":
            record_registry_event(
                registry="input_reader",
                name=normalized_format,
                action="keep",
                policy=policy,
                existing_source_module=existing.source_module,
                new_source_module=source,
                detail="kept existing registration",
            )
            return existing
        action = "replace_with_warning" if policy == "warn" else "replace"
        record_registry_event(
            registry="input_reader",
            name=normalized_format,
            action=action,
            policy=policy,
            existing_source_module=existing.source_module,
            new_source_module=source,
            detail="replaced existing registration",
        )
        unregister_input_reader(normalized_format)

    ext_conflicts = _extension_conflicts(normalized_format, normalized_extensions)
    if ext_conflicts:
        first_conflict = next(iter(ext_conflicts.values()))[0]
        if policy == "error":
            for conflicting_registration, captured_extensions in ext_conflicts.values():
                for ext in sorted(captured_extensions):
                    record_registry_event(
                        registry="input_reader_extension",
                        name=ext,
                        action="error",
                        policy=policy,
                        existing_source_module=conflicting_registration.source_module,
                        new_source_module=source,
                        detail=(
                            f"extension {ext} already registered to "
                            f"{conflicting_registration.input_format}"
                        ),
                    )
            raise RuntimeError(
                "input reader extensions already registered: "
                + ", ".join(
                    f"{ext}->{registration.input_format}"
                    for registration, captured_extensions in ext_conflicts.values()
                    for ext in sorted(captured_extensions)
                )
            )
        if policy == "keep":
            for conflicting_registration, captured_extensions in ext_conflicts.values():
                for ext in sorted(captured_extensions):
                    record_registry_event(
                        registry="input_reader_extension",
                        name=ext,
                        action="keep",
                        policy=policy,
                        existing_source_module=conflicting_registration.source_module,
                        new_source_module=source,
                        detail=(
                            f"kept existing extension owner "
                            f"{conflicting_registration.input_format}"
                        ),
                    )
            return first_conflict

        action = "replace_with_warning" if policy == "warn" else "replace"
        for conflicting_registration, captured_extensions in ext_conflicts.values():
            remaining_extensions = tuple(
                ext for ext in conflicting_registration.extensions if ext not in captured_extensions
            )
            if remaining_extensions:
                state.input_readers_by_format[conflicting_registration.input_format] = replace(
                    conflicting_registration,
                    extensions=remaining_extensions,
                )
            else:
                state.input_readers_by_format.pop(conflicting_registration.input_format, None)
            for ext in sorted(captured_extensions):
                record_registry_event(
                    registry="input_reader_extension",
                    name=ext,
                    action=action,
                    policy=policy,
                    existing_source_module=conflicting_registration.source_module,
                    new_source_module=source,
                    detail=(
                        f"extension {ext} moved from {conflicting_registration.input_format} "
                        f"to {normalized_format}"
                    ),
                )
                state.input_readers_by_extension.pop(ext, None)

    registration = InputReaderRegistration(
        input_format=normalized_format,
        extensions=normalized_extensions,
        loader=loader,
        domain=normalized_domain,
        domain_metadata=normalized_domain_metadata,
        source_module=source,
    )
    state.input_readers_by_format[normalized_format] = registration
    for ext in normalized_extensions:
        state.input_readers_by_extension[ext] = registration
    return registration


def register_input_reader(
    input_format: str,
    extensions: Iterable[str],
    loader: InputReaderFn,
    *,
    domain: Optional[str] = None,
    domain_metadata: Optional[Mapping[str, Any]] = None,
    source_module: Optional[str] = None,
    on_conflict: Optional[str] = None,
) -> InputReaderRegistration:
    _ensure_builtin_input_readers()
    effective_source = source_module or infer_caller_module()
    return _register_input_reader(
        input_format,
        extensions,
        loader,
        domain=domain,
        domain_metadata=domain_metadata,
        source_module=effective_source,
        on_conflict=on_conflict,
    )


def unregister_input_reader(input_format: str) -> Optional[InputReaderRegistration]:
    _ensure_builtin_input_readers()
    state = get_runtime_state()
    normalized_format = _normalize_input_format(input_format)
    registration = state.input_readers_by_format.pop(normalized_format, None)
    if registration is None:
        return None
    for ext, current in list(state.input_readers_by_extension.items()):
        if current.input_format == normalized_format:
            del state.input_readers_by_extension[ext]
    return registration


def get_input_reader(path: str) -> InputReaderRegistration:
    _ensure_builtin_input_readers()
    ext = _ext(path)
    registration = get_runtime_state().input_readers_by_extension.get(ext)
    if registration is None:
        raise RuntimeError(f"unsupported input file type: {path}")
    return registration


def list_input_formats() -> List[str]:
    _ensure_builtin_input_readers()
    return sorted(get_runtime_state().input_readers_by_format.keys())


def list_input_reader_details() -> List[Dict[str, Any]]:
    _ensure_builtin_input_readers()
    state = get_runtime_state()
    return [
        {
            "input_format": registration.input_format,
            "extensions": list(registration.extensions),
            "domain": registration.domain,
            "domain_metadata": dict(registration.domain_metadata),
            "source_module": registration.source_module,
        }
        for registration in sorted(state.input_readers_by_format.values(), key=lambda item: item.input_format)
    ]


def list_input_reader_domains() -> List[Dict[str, Any]]:
    _ensure_builtin_input_readers()
    state = get_runtime_state()
    return summarize_registry_domains(
        sorted(state.input_readers_by_format.values(), key=lambda item: item.input_format),
        item_name_attr="input_format",
        list_key="input_formats",
    )


def _resolve_tesseract_cmd() -> Optional[str]:
    env_cmd = os.environ.get("TESSERACT_CMD")
    if env_cmd and os.path.exists(env_cmd):
        return env_cmd
    candidates = [
        r"C:\Program Files\Tesseract-OCR\tesseract.exe",
        r"C:\Program Files (x86)\Tesseract-OCR\tesseract.exe",
        "/usr/bin/tesseract",
        "/usr/local/bin/tesseract",
        "/opt/homebrew/bin/tesseract",
    ]
    for c in candidates:
        if os.path.exists(c):
            return c
    return None


def _ocr_preprocess_mode(value: Optional[str]) -> str:
    mode = str(value or os.environ.get("AIWF_OCR_PREPROCESS") or "adaptive").strip().lower()
    if mode not in {"none", "off", "gray", "adaptive"}:
        return "adaptive"
    return mode


def _ocr_try_modes(value: Optional[str]) -> List[str]:
    raw = str(value or os.environ.get("AIWF_OCR_TRY_MODES") or "").strip()
    if raw:
        out: List[str] = []
        for token in raw.split(","):
            m = _ocr_preprocess_mode(token)
            if m not in out:
                out.append(m)
        if out:
            return out
    mode = _ocr_preprocess_mode(value)
    if mode in {"none", "off"}:
        return ["none"]
    if mode == "gray":
        return ["gray", "none"]
    return ["adaptive", "gray", "none"]


def _preprocess_image_for_ocr(image: Any, mode: str) -> Any:
    if mode in {"none", "off"}:
        return image
    try:
        from PIL import ImageFilter, ImageOps  # type: ignore
    except Exception:
        return image
    gray = ImageOps.grayscale(image)
    if mode == "gray":
        return gray
    enhanced = ImageOps.autocontrast(gray)
    filtered = enhanced.filter(ImageFilter.MedianFilter(size=3))
    return filtered.point(lambda x: 255 if x > 160 else 0, mode="1")


def _ocr_text_score(text: str) -> int:
    if not text:
        return 0
    s = str(text).strip()
    if not s:
        return 0
    return len(re.findall(r"[A-Za-z0-9\u4e00-\u9fff]", s))


def _ocr_extract_text(pytesseract: Any, image: Any, lang: str, config: str, modes: List[str]) -> str:
    best = ""
    best_score = -1
    successful_modes = 0
    last_err: Optional[Exception] = None
    for mode in modes:
        try:
            processed = _preprocess_image_for_ocr(image, mode)
            text = pytesseract.image_to_string(processed, lang=lang, config=config)
        except Exception as exc:
            last_err = exc
            continue
        successful_modes += 1
        score = _ocr_text_score(text)
        if score > best_score:
            best = text
            best_score = score
    if successful_modes == 0 and last_err is not None:
        raise RuntimeError(f"OCR failed for all preprocess modes: {last_err}") from last_err
    return best


def _split_text_to_rows(text: str, path: str, source_type: str, by_line: bool = False) -> List[Dict[str, Any]]:
    if by_line:
        chunks = [ln.strip() for ln in text.splitlines() if ln.strip()]
    else:
        chunks = [p.strip() for p in text.split("\n\n") if p.strip()]
    rows: List[Dict[str, Any]] = []
    for i, c in enumerate(chunks):
        rows.append(
            {
                "text": c,
                "source_file": os.path.basename(path),
                "source_path": path,
                "source_type": source_type,
                "chunk_index": i,
            }
        )
    return rows


def _read_text_with_fallback(path: str) -> str:
    encodings = ["utf-8-sig", "utf-8", "gb18030", "gbk"]
    last_err: Optional[Exception] = None
    for enc in encodings:
        try:
            with open(path, "r", encoding=enc, errors="strict") as f:
                return f.read()
        except Exception as e:
            last_err = e
    with open(path, "r", encoding="utf-8", errors="ignore") as f:
        txt = f.read()
    if txt:
        return txt
    raise RuntimeError(f"cannot decode text file: {path}, last_err={last_err}")


def read_txt(path: str, *, by_line: bool = False) -> List[Dict[str, Any]]:
    text = _read_text_with_fallback(path)
    return _split_text_to_rows(text, path, "txt", by_line=by_line)


def read_docx(path: str, *, by_line: bool = False) -> List[Dict[str, Any]]:
    try:
        from docx import Document  # type: ignore
    except Exception as e:
        raise RuntimeError(f"docx support requires python-docx: {e}")
    doc = Document(path)
    text = "\n".join(p.text for p in doc.paragraphs if p.text and p.text.strip())
    return _split_text_to_rows(text, path, "docx", by_line=by_line)


def read_pdf(path: str, *, by_line: bool = False) -> List[Dict[str, Any]]:
    try:
        from pypdf import PdfReader  # type: ignore
    except Exception as e:
        raise RuntimeError(f"pdf support requires pypdf: {e}")
    reader = PdfReader(path)
    rows: List[Dict[str, Any]] = []
    for page_idx, page in enumerate(reader.pages):
        text = page.extract_text() or ""
        chunks = _split_text_to_rows(text, path, "pdf", by_line=by_line)
        for c in chunks:
            c["page"] = page_idx + 1
        rows.extend(chunks)
    return rows


def read_image(
    path: str,
    *,
    by_line: bool = False,
    ocr_lang: Optional[str] = None,
    ocr_config: Optional[str] = None,
    ocr_preprocess: Optional[str] = None,
) -> List[Dict[str, Any]]:
    try:
        from PIL import Image  # type: ignore
    except Exception as e:
        raise RuntimeError(f"image support requires Pillow: {e}")
    try:
        import pytesseract  # type: ignore
    except Exception as e:
        raise RuntimeError(f"image OCR support requires pytesseract: {e}")
    cmd = _resolve_tesseract_cmd()
    if cmd:
        pytesseract.pytesseract.tesseract_cmd = cmd
    lang = str(ocr_lang or os.environ.get("AIWF_OCR_LANG") or "eng+chi_sim").strip()
    config = str(ocr_config or os.environ.get("AIWF_OCR_CONFIG") or "--oem 1 --psm 6").strip()
    modes = _ocr_try_modes(ocr_preprocess)
    with Image.open(path) as image:
        text = _ocr_extract_text(pytesseract, image, lang, config, modes)
    return _split_text_to_rows(text, path, "image", by_line=by_line)


def read_xlsx(path: str, *, include_all_sheets: bool = False) -> List[Dict[str, Any]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as e:
        raise RuntimeError(f"xlsx support requires openpyxl: {e}")
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        out: List[Dict[str, Any]] = []
        sheet_names = wb.sheetnames if include_all_sheets else [wb.sheetnames[0]]
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            try:
                header_raw = next(rows_iter)
            except StopIteration:
                continue
            headers = [str(h).strip() if h is not None and str(h).strip() else f"col_{i+1}" for i, h in enumerate(header_raw)]
            for row_idx, row in enumerate(rows_iter, start=2):
                obj: Dict[str, Any] = {}
                non_empty = False
                for i, v in enumerate(row):
                    obj[headers[i]] = v
                    if v is not None and str(v).strip() != "":
                        non_empty = True
                if non_empty:
                    obj["source_file"] = os.path.basename(path)
                    obj["source_path"] = path
                    obj["source_type"] = "xlsx"
                    obj["sheet_name"] = sheet_name
                    obj["row_index"] = row_idx
                    out.append(obj)
        return out
    finally:
        wb.close()


def _load_txt_input(path: str, options: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    return read_txt(path, by_line=bool(options.get("text_by_line", False))), {"input_format": "txt"}


def _load_docx_input(path: str, options: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    return read_docx(path, by_line=bool(options.get("text_by_line", False))), {"input_format": "docx"}


def _load_pdf_input(path: str, options: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    return read_pdf(path, by_line=bool(options.get("text_by_line", False))), {"input_format": "pdf"}


def _load_image_input(path: str, options: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    if not bool(options.get("ocr_enabled", True)):
        return [], {"input_format": "image", "skipped": True, "reason": "ocr disabled"}
    return (
        read_image(
            path,
            by_line=bool(options.get("text_by_line", False)),
            ocr_lang=options.get("ocr_lang"),
            ocr_config=options.get("ocr_config"),
            ocr_preprocess=options.get("ocr_preprocess"),
        ),
        {"input_format": "image"},
    )


def _load_xlsx_input(path: str, options: Dict[str, Any]) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    return (
        read_xlsx(path, include_all_sheets=bool(options.get("xlsx_all_sheets", False))),
        {"input_format": "xlsx"},
    )


def load_rows_from_file(
    path: str,
    *,
    text_by_line: bool = False,
    ocr_enabled: bool = True,
    ocr_lang: Optional[str] = None,
    ocr_config: Optional[str] = None,
    ocr_preprocess: Optional[str] = None,
    xlsx_all_sheets: bool = False,
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    _ensure_builtin_input_readers()
    registration = get_input_reader(path)
    return registration.loader(
        path,
        {
            "text_by_line": text_by_line,
            "ocr_enabled": ocr_enabled,
            "ocr_lang": ocr_lang,
            "ocr_config": ocr_config,
            "ocr_preprocess": ocr_preprocess,
            "xlsx_all_sheets": xlsx_all_sheets,
        },
    )


def load_rows_from_files(
    paths: List[str],
    *,
    text_by_line: bool = False,
    ocr_enabled: bool = True,
    ocr_lang: Optional[str] = None,
    ocr_config: Optional[str] = None,
    ocr_preprocess: Optional[str] = None,
    xlsx_all_sheets: bool = False,
    max_retries: int = 0,
    on_file_error: str = "skip",
) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    _ensure_builtin_input_readers()
    all_rows: List[Dict[str, Any]] = []
    formats: List[str] = []
    skipped_files: List[Dict[str, Any]] = []
    failed_files: List[Dict[str, Any]] = []
    for p in paths:
        last_err: Optional[str] = None
        loaded = False
        for _ in range(max_retries + 1):
            try:
                rows, meta = load_rows_from_file(
                    p,
                    text_by_line=text_by_line,
                    ocr_enabled=ocr_enabled,
                    ocr_lang=ocr_lang,
                    ocr_config=ocr_config,
                    ocr_preprocess=ocr_preprocess,
                    xlsx_all_sheets=xlsx_all_sheets,
                )
                fmt = str(meta.get("input_format"))
                formats.append(fmt)
                if meta.get("skipped"):
                    skipped_files.append({"path": p, "reason": str(meta.get("reason") or "skipped")})
                else:
                    all_rows.extend(rows)
                loaded = True
                break
            except Exception as e:
                last_err = str(e)
        if not loaded:
            failed_files.append({"path": p, "error": last_err or "unknown error"})
            if on_file_error == "raise":
                raise RuntimeError(f"failed to load file {p}: {last_err}")
    return all_rows, {
        "input_format": ",".join(formats),
        "file_count": len(paths),
        "skipped_files": skipped_files,
        "failed_files": failed_files,
    }
