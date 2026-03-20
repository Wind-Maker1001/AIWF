from __future__ import annotations

import os
from typing import Any, Dict, List, Optional

from aiwf.office_style import (
    office_font_name,
    office_is_high_quality,
    office_layout_settings,
    office_quality_mode,
    office_text,
    office_theme_settings,
)


def write_fin_xlsx(
    xlsx_path: str,
    rows: List[Dict[str, Any]],
    image_path: Optional[str] = None,
    params: Optional[Dict[str, Any]] = None,
    *,
    to_decimal,
    build_profile,
    utc_now_str,
) -> None:
    from openpyxl import Workbook  # type: ignore
    from openpyxl.styles import Alignment, Font, PatternFill  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    wb = Workbook()
    ws = wb.active
    ws.title = "detail"
    if rows:
        columns = list(rows[0].keys())
        seen = set(columns)
        for row in rows[1:]:
            for key in row.keys():
                if key not in seen:
                    columns.append(key)
                    seen.add(key)
    else:
        columns = ["id", "amount"]

    theme = office_theme_settings(params)
    layout = office_layout_settings(params)
    high_quality = office_is_high_quality(params)
    font_name = office_font_name(params)
    header_fill = PatternFill(fill_type="solid", fgColor=str(theme.get("primary_hex", "1F4E78")))
    alt_fill = PatternFill(fill_type="solid", fgColor="F4F7FB")
    header_font = Font(name=font_name, color="FFFFFF", bold=True)
    body_font = Font(name=font_name)
    title_font = Font(name=font_name, bold=True, color=str(theme.get("primary_hex", "1F4E78")), size=14)
    center = Alignment(horizontal="center", vertical="center")
    left = Alignment(horizontal="left", vertical="center")

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(2, len(columns)))
    ws.cell(row=1, column=1, value=str(theme.get("report_title"))).font = title_font
    ws.cell(row=1, column=1).alignment = left
    ws.row_dimensions[1].height = 24

    header_row = 2
    for col_idx, column in enumerate(columns, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
    for row_idx, row in enumerate(rows, start=header_row + 1):
        for col_idx, column in enumerate(columns, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row.get(column))
            cell.font = body_font
            cell.alignment = left
            if high_quality and row_idx % 2 == 1:
                cell.fill = alt_fill

    ws.freeze_panes = "A3"
    if columns:
        ws.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(len(columns))}{max(header_row + 1, len(rows) + header_row)}"
        )

    numeric_fields = []
    for column in columns:
        if column.lower() in {"amount", "sum", "min", "max", "avg"}:
            numeric_fields.append(column)
            continue
        for row in rows[:100]:
            if to_decimal(row.get(column)) is not None:
                numeric_fields.append(column)
                break
    for column in numeric_fields:
        col_idx = columns.index(column) + 1
        for row_idx in range(header_row + 1, len(rows) + header_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = "#,##0.00"

    for col_idx, column in enumerate(columns, start=1):
        max_len = len(str(column))
        for row in rows[:500]:
            max_len = max(max_len, len(str(row.get(column) if row.get(column) is not None else "")))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 48)

    profile_like = build_profile(rows, {"input_rows": len(rows), "output_rows": len(rows)}, "xlsx.export")
    summary = wb.create_sheet("summary")
    summary["A1"] = office_text("指标", "Metric", params)
    summary["B1"] = office_text("数值", "Value", params)
    summary["A1"].fill = header_fill
    summary["B1"].fill = header_fill
    summary["A1"].font = header_font
    summary["B1"].font = header_font
    summary["A1"].alignment = center
    summary["B1"].alignment = center
    metrics = [
        (office_text("主题", "Theme", params), theme.get("display_name", theme.get("name"))),
        (office_text("质量模式", "Quality Mode", params), office_quality_mode(params)),
        (office_text("行数", "Rows", params), profile_like.get("rows")),
        (office_text("列数", "Columns", params), len(columns)),
        (office_text("金额总计", "Sum Amount", params), profile_like.get("sum_amount")),
        (office_text("金额最小值", "Min Amount", params), profile_like.get("min_amount")),
        (office_text("金额最大值", "Max Amount", params), profile_like.get("max_amount")),
        (office_text("金额均值", "Avg Amount", params), profile_like.get("avg_amount")),
        (office_text("生成时间", "Generated At", params), utc_now_str()),
    ]
    for i, (key, value) in enumerate(metrics, start=2):
        left_cell = summary.cell(row=i, column=1, value=key)
        right_cell = summary.cell(row=i, column=2, value=value)
        left_cell.font = body_font
        right_cell.font = body_font
        left_cell.alignment = left
        right_cell.alignment = left
        if high_quality and i % 2 == 0:
            left_cell.fill = alt_fill
            right_cell.fill = alt_fill
    summary.column_dimensions["A"].width = 24
    summary.column_dimensions["B"].width = 28
    if image_path and os.path.isfile(image_path):
        try:
            from openpyxl.drawing.image import Image as XLImage  # type: ignore

            summary.add_image(XLImage(image_path), "D2")
        except Exception:
            pass

    _ = layout
    wb.save(xlsx_path)
