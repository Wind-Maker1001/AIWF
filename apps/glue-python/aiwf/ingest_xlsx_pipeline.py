from __future__ import annotations

import os
import re
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from aiwf.ingest_docling_pipeline import coerce_extraction_result, extract_with_docling
from aiwf.quality_contract import (
    build_xlsx_quality_report,
    canonicalize_header,
    normalize_value_for_field,
)


def _load_calamine_rows(path: str) -> tuple[Optional[dict[str, list[list[Any]]]], str]:
    try:
        from python_calamine import load_workbook  # type: ignore
    except Exception:
        return None, "openpyxl"

    try:
        workbook = load_workbook(path)
    except Exception:
        return None, "openpyxl"

    sheets: dict[str, list[list[Any]]] = {}
    sheet_names = []
    try:
        if hasattr(workbook, "sheet_names"):
            sheet_names = list(workbook.sheet_names)
        elif hasattr(workbook, "sheets_metadata"):
            sheet_names = [str(item.name) for item in workbook.sheets_metadata]
    except Exception:
        sheet_names = []

    for name in sheet_names:
        try:
            sheet = workbook.get_sheet_by_name(name)
        except Exception:
            continue
        rows: list[list[Any]] = []
        try:
            for row in sheet.to_python():
                rows.append(list(row))
        except Exception:
            continue
        sheets[str(name)] = rows
    if not sheets:
        return None, "openpyxl"
    return sheets, "calamine+openpyxl"


def _merged_value(ws: Any, row_idx: int, col_idx: int, merged_map: dict[tuple[int, int], Any]) -> Any:
    if (row_idx, col_idx) in merged_map:
        return merged_map[(row_idx, col_idx)]
    return ws.cell(row=row_idx, column=col_idx).value


def _build_merged_map(ws: Any) -> dict[tuple[int, int], Any]:
    merged_map: dict[tuple[int, int], Any] = {}
    try:
        ranges = list(ws.merged_cells.ranges)
    except Exception:
        return merged_map
    for merged_range in ranges:
        min_col = int(getattr(merged_range, "min_col", 0) or 0)
        max_col = int(getattr(merged_range, "max_col", 0) or 0)
        min_row = int(getattr(merged_range, "min_row", 0) or 0)
        max_row = int(getattr(merged_range, "max_row", 0) or 0)
        if min_col <= 0 or min_row <= 0:
            continue
        value = ws.cell(row=min_row, column=min_col).value
        for row_idx in range(min_row, max_row + 1):
            for col_idx in range(min_col, max_col + 1):
                merged_map[(row_idx, col_idx)] = value
    return merged_map


def _normalize_header_piece(value: Any) -> str:
    text = str(value or "").strip()
    return text


def _header_looks_like_data(value: str) -> bool:
    text = str(value or "").strip()
    if not text:
        return False
    compact = re.sub(r"\s+", "", text)
    if re.fullmatch(r"[0-9,.\-+/年月日:（）()]+", compact):
        return True
    if re.fullmatch(r"[A-Za-z0-9\-_/.]+", compact) and any(ch.isdigit() for ch in compact):
        return True
    return False


def _candidate_header(
    ws: Any,
    *,
    merged_map: dict[tuple[int, int], Any],
    spec: dict[str, Any],
    max_scan_rows: int = 5,
    max_header_rows: int = 2,
) -> tuple[int, int, list[str], list[float], list[str]]:
    max_row = max(1, min(int(getattr(ws, "max_row", 1) or 1), max_scan_rows))
    max_col = max(1, int(getattr(ws, "max_column", 1) or 1))
    best_score = -1.0
    best_start = 1
    best_end = 1
    best_headers: list[str] = []
    best_confidences: list[float] = []
    best_raw_headers: list[str] = []

    for start_row in range(1, min(3, max_row) + 1):
        for end_row in range(start_row, min(max_row, start_row + max_header_rows - 1) + 1):
            folded: list[str] = []
            confidences: list[float] = []
            raw_headers: list[str] = []
            recognized = 0
            placeholders = 0
            data_like = 0
            start_row_data_like = 0
            unique = set()
            for col_idx in range(1, max_col + 1):
                parts = []
                for row_idx in range(start_row, end_row + 1):
                    part = _normalize_header_piece(_merged_value(ws, row_idx, col_idx, merged_map))
                    if part:
                        parts.append(part)
                raw_header = " ".join(parts).strip() or f"col_{col_idx}"
                canonical, confidence, _matched = canonicalize_header(raw_header, spec)
                folded.append(canonical or f"col_{col_idx}")
                confidences.append(confidence)
                raw_headers.append(raw_header)
                if canonical.startswith("col_"):
                    placeholders += 1
                if confidence >= 0.85:
                    recognized += 1
                if _header_looks_like_data(raw_header):
                    data_like += 1
                start_raw = _normalize_header_piece(_merged_value(ws, start_row, col_idx, merged_map))
                if _header_looks_like_data(start_raw):
                    start_row_data_like += 1
                unique.add(canonical)
            avg_confidence = (sum(confidences) / len(confidences)) if confidences else 0.0
            score = (
                (len(unique) / max_col) * 0.4
                + (recognized / max_col) * 0.8
                + avg_confidence * 0.8
                - (placeholders / max_col) * 0.4
                - (data_like / max_col) * 0.9
            )
            if start_row > 1 and (start_row_data_like / max_col) >= 0.5:
                score -= 1.5
            if score > best_score:
                best_score = score
                best_start = start_row
                best_end = end_row
                best_headers = folded
                best_confidences = confidences
                best_raw_headers = raw_headers
    return best_start, best_end, best_headers, best_confidences, best_raw_headers


def _is_blank_row(values: Sequence[Any]) -> bool:
    return all(value is None or str(value).strip() == "" for value in values)


def _iter_sheet_rows(ws: Any, calamine_rows: Optional[list[list[Any]]]) -> Iterable[list[Any]]:
    if isinstance(calamine_rows, list) and calamine_rows:
        for row in calamine_rows:
            yield list(row)
        return
    for row in ws.iter_rows(values_only=True):
        yield list(row)


def extract_xlsx_rows(
    path: str,
    *,
    include_all_sheets: bool = True,
    spec: Optional[Dict[str, Any]] = None,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    spec_obj = dict(spec or {})

    docling = coerce_extraction_result(extract_with_docling(path))
    if docling is not None and docling.rows:
        rows = list(docling.rows or [])
        sheet_frames = list(docling.sheet_frames or [])
        table_cells = list(docling.table_cells or [])
        quality_report = build_xlsx_quality_report(rows, sheet_frames, spec_obj)
        return rows, {
            "input_format": "xlsx",
            "engine": "docling",
            "sheet_frames": sheet_frames,
            "table_cells": table_cells,
            "quality_report": quality_report,
            "quality_metrics": quality_report.get("metrics") if isinstance(quality_report.get("metrics"), dict) else {},
            "engine_trace": list(docling.engine_trace or []),
            "quality_blocked": bool(quality_report.get("blocked")),
            "quality_error": "; ".join(quality_report.get("errors") or []),
        }

    try:
        from openpyxl import load_workbook  # type: ignore
    except Exception as exc:
        raise RuntimeError(f"xlsx support requires openpyxl: {exc}")

    calamine_rows_by_sheet, engine = _load_calamine_rows(path)
    workbook_formula = load_workbook(path, read_only=False, data_only=False)
    workbook_values = load_workbook(path, read_only=True, data_only=True)
    try:
        allowlist = spec_obj.get("sheet_allowlist")
        allowed_sheets = {str(item).strip() for item in allowlist} if isinstance(allowlist, list) else set()
        include_hidden = bool(spec_obj.get("include_hidden_sheets", False))
        selected_sheet_names: list[str] = []
        for index, name in enumerate(workbook_values.sheetnames):
            if allowed_sheets and name not in allowed_sheets:
                continue
            if not include_all_sheets and index > 0 and not allowed_sheets:
                continue
            try:
                hidden_state = str(getattr(workbook_formula[name], "sheet_state", "visible") or "visible").lower()
            except Exception:
                hidden_state = "visible"
            if hidden_state != "visible" and not include_hidden:
                continue
            selected_sheet_names.append(name)

        all_rows: list[dict[str, Any]] = []
        sheet_frames: list[dict[str, Any]] = []
        table_cells: list[dict[str, Any]] = []
        workbook_name = os.path.basename(path)
        for sheet_index, sheet_name in enumerate(selected_sheet_names):
            ws_values = workbook_values[sheet_name]
            ws_formula = workbook_formula[sheet_name]
            merged_map = _build_merged_map(ws_formula)
            header_start, header_end, headers, header_confidences, raw_headers = _candidate_header(
                ws_formula,
                merged_map=merged_map,
                spec=spec_obj,
            )
            frame_rows: list[dict[str, Any]] = []
            blank_rows = 0
            numeric_total = 0
            numeric_parsed = 0
            date_total = 0
            date_parsed = 0
            formula_cells = 0
            formula_mismatches = 0
            rows_iter = list(_iter_sheet_rows(ws_values, (calamine_rows_by_sheet or {}).get(sheet_name)))
            for row_offset, raw_row in enumerate(rows_iter, start=1):
                if row_offset <= header_end:
                    continue
                padded = list(raw_row) + [None] * max(0, len(headers) - len(raw_row))
                values = padded[: len(headers)]
                if _is_blank_row(values):
                    blank_rows += 1
                    continue
                item: dict[str, Any] = {
                    "source_file": workbook_name,
                    "source_path": path,
                    "source_type": "xlsx",
                    "workbook_name": workbook_name,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "row_index": row_offset,
                }
                for col_index, header in enumerate(headers, start=1):
                    value = values[col_index - 1] if col_index - 1 < len(values) else None
                    normalized_value = normalize_value_for_field(
                        value,
                        header,
                        raw_header=raw_headers[col_index - 1] if col_index - 1 < len(raw_headers) else header,
                    )
                    item[header] = normalized_value
                    table_cells.append(
                        {
                            "cell_id": f"{sheet_name}_{row_offset}_{col_index}",
                            "row": row_offset,
                            "col": col_index,
                            "text": "" if normalized_value is None else str(normalized_value),
                            "bbox": [0, 0, 0, 0],
                            "sheet_name": sheet_name,
                            "source_path": path,
                        }
                    )
                    if any(token in header for token in ("amount", "amt", "score", "id")) and value not in {None, ""}:
                        numeric_total += 1
                        if normalized_value not in {"", None}:
                            try:
                                float(str(normalized_value).replace(",", ""))
                                numeric_parsed += 1
                            except Exception:
                                pass
                    if "date" in header or header.endswith("_at"):
                        if value not in {None, ""}:
                            date_total += 1
                            if normalized_value not in {"", None} and str(normalized_value) != str(value):
                                date_parsed += 1

                    formula_cell = ws_formula.cell(row=row_offset, column=col_index)
                    if isinstance(formula_cell.value, str) and formula_cell.value.startswith("="):
                        formula_cells += 1
                        if value in {None, ""}:
                            formula_mismatches += 1
                frame_rows.append(item)
                all_rows.append(item)

            try:
                table_names = list((ws_formula.tables or {}).keys())
            except Exception:
                table_names = []
            sheet_frames.append(
                {
                    "workbook_name": workbook_name,
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "header_row_span": [header_start, header_end],
                    "header_confidence": round(
                        (sum(header_confidences) / len(header_confidences)) if header_confidences else 0.0,
                        6,
                    ),
                    "header_labels": raw_headers,
                    "table_name": table_names[0] if table_names else "",
                    "columns": headers,
                    "row_count": len(frame_rows),
                    "blank_rows": blank_rows,
                    "numeric_cells_total": numeric_total,
                    "numeric_cells_parsed": numeric_parsed,
                    "date_cells_total": date_total,
                    "date_cells_parsed": date_parsed,
                    "formula_cells": formula_cells,
                    "formula_mismatches": formula_mismatches,
                    "hidden": str(getattr(ws_formula, "sheet_state", "visible") or "visible").lower() != "visible",
                    "source_path": path,
                }
            )

        quality_report = build_xlsx_quality_report(all_rows, sheet_frames, spec_obj)
        meta = {
            "input_format": "xlsx",
            "engine": engine,
            "sheet_frames": sheet_frames,
            "table_cells": table_cells,
            "quality_report": quality_report,
            "quality_metrics": quality_report.get("metrics") if isinstance(quality_report.get("metrics"), dict) else {},
            "engine_trace": [{"engine": engine, "ok": True, "sheet_count": len(sheet_frames)}],
            "quality_blocked": bool(quality_report.get("blocked")),
            "quality_error": "; ".join(quality_report.get("errors") or []),
        }
        return all_rows, meta
    finally:
        workbook_values.close()
        workbook_formula.close()
