import os
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

from aiwf import ingest
from aiwf.ingest_file_readers import load_pdf_input
from aiwf.ingest_image_pipeline import extract_image_rows
from aiwf.ingest_xlsx_pipeline import extract_xlsx_rows


class IngestTests(unittest.TestCase):
    def test_ocr_try_modes_defaults_and_parse(self):
        with patch.dict(os.environ, {}, clear=True):
            self.assertEqual(ingest._ocr_try_modes(None), ["adaptive", "gray", "none"])
        with patch.dict(os.environ, {"AIWF_OCR_TRY_MODES": "gray,none"}, clear=False):
            self.assertEqual(ingest._ocr_try_modes(None), ["gray", "none"])
        self.assertEqual(ingest._ocr_try_modes("none"), ["none"])

    def test_ocr_extract_text_prefers_higher_signal_score(self):
        class DummyT:
            @staticmethod
            def image_to_string(img, lang=None, config=None):
                mapping = {
                    "adaptive": "abc",
                    "gray": "abc123中文",
                    "none": "ab",
                }
                return mapping.get(str(img), "")

        with patch("aiwf.ingest._preprocess_image_for_ocr") as pre:
            pre.side_effect = lambda image, mode: mode
            text = ingest._ocr_extract_text(DummyT(), object(), "eng+chi_sim", "--psm 6", ["adaptive", "gray", "none"])
            self.assertEqual(text, "abc123中文")

    def test_ocr_extract_text_raises_when_all_modes_fail(self):
        class DummyT:
            @staticmethod
            def image_to_string(img, lang=None, config=None):
                raise RuntimeError(f"ocr failed for {img}")

        with patch("aiwf.ingest._preprocess_image_for_ocr") as pre:
            pre.side_effect = lambda image, mode: mode
            with self.assertRaisesRegex(RuntimeError, "OCR failed for all preprocess modes"):
                ingest._ocr_extract_text(DummyT(), object(), "eng+chi_sim", "--psm 6", ["adaptive", "gray"])

    def test_resolve_tesseract_cmd_prefers_env(self):
        with patch.dict(os.environ, {"TESSERACT_CMD": r"C:\custom\tesseract.exe"}, clear=False):
            with patch("aiwf.ingest.os.path.exists") as exists:
                exists.side_effect = lambda p: p == r"C:\custom\tesseract.exe"
                self.assertEqual(ingest._resolve_tesseract_cmd(), r"C:\custom\tesseract.exe")

    def test_resolve_tesseract_cmd_uses_known_candidates(self):
        with patch.dict(os.environ, {}, clear=True):
            with patch("aiwf.ingest.os.path.exists") as exists:
                exists.side_effect = lambda p: p == r"C:\Program Files\Tesseract-OCR\tesseract.exe"
                self.assertEqual(ingest._resolve_tesseract_cmd(), r"C:\Program Files\Tesseract-OCR\tesseract.exe")

    def test_read_txt(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.txt")
            with open(p, "w", encoding="utf-8") as f:
                f.write("line1\n\nline2")
            rows, meta = ingest.load_rows_from_file(p)
            self.assertEqual(meta["input_format"], "txt")
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["text"], "line1")

    def test_read_xlsx(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.xlsx")
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "S1"
            ws.append(["id", "amount"])
            ws.append([1, 10.5])
            ws2 = wb.create_sheet("S2")
            ws2.append(["id", "amount"])
            ws2.append([2, 20.0])
            wb.save(p)

            rows, meta = ingest.load_rows_from_file(p)
            self.assertEqual(meta["input_format"], "xlsx")
            self.assertEqual(len(rows), 2)
            self.assertEqual(rows[0]["id"], 1)
            self.assertEqual(rows[0]["source_type"], "xlsx")
            self.assertEqual(rows[0]["sheet_name"], "S1")
            self.assertEqual({r["sheet_name"] for r in rows}, {"S1", "S2"})

            rows_single, _ = ingest.load_rows_from_file(p, xlsx_all_sheets=False)
            self.assertEqual(len(rows_single), 1)
            self.assertEqual({r["sheet_name"] for r in rows_single}, {"S1"})

    def test_read_xlsx_reports_quality_block_when_required_columns_missing(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.xlsx")
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "S1"
            ws.append(["id", "note"])
            ws.append([1, "hello"])
            wb.save(p)

            rows, meta = ingest.load_rows_from_file(
                p,
                extra_options={
                    "xlsx_rules": {
                        "required_columns": ["id", "amount"],
                        "sheet_row_count_min": 1,
                    }
                },
            )
            self.assertEqual(len(rows), 1)
            self.assertTrue(meta["quality_blocked"])
            self.assertIn("required", str(meta["quality_error"]))

    def test_extract_image_rows_prefers_structured_blocks(self):
        with patch("aiwf.ingest_image_pipeline._extract_with_paddleocr") as extract_paddle:
            extract_paddle.return_value = (
                [
                    {
                        "block_id": "img_blk_0001",
                        "block_type": "text",
                        "bbox": [1, 2, 30, 40],
                        "text": "Tax policy should be supported.",
                        "confidence": 0.98,
                        "line_no": 1,
                        "page_no": 1,
                        "source_path": "demo.png",
                    }
                ],
                "paddleocr",
            )
            rows, meta = extract_image_rows(
                "demo.png",
                spec={"image_rules": {"ocr_confidence_avg_min": 0.8}},
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["image_block_id"], "img_blk_0001")
        self.assertEqual(meta["ocr_engine"], "paddleocr")
        self.assertFalse(meta["quality_blocked"])
        self.assertIn("engine_trace", meta)
        self.assertIn("quality_metrics", meta)

    def test_extract_image_rows_recovers_table_cells_from_layout(self):
        with patch("aiwf.ingest_image_pipeline._extract_with_paddleocr") as extract_paddle:
            extract_paddle.return_value = (
                [
                    {"block_id": "b1", "block_type": "text", "bbox": [10, 10, 80, 40], "text": "编号", "confidence": 0.95, "line_no": 1, "page_no": 1, "source_path": "demo.png"},
                    {"block_id": "b2", "block_type": "text", "bbox": [120, 10, 230, 40], "text": "金额(万元)", "confidence": 0.95, "line_no": 2, "page_no": 1, "source_path": "demo.png"},
                    {"block_id": "b3", "block_type": "text", "bbox": [10, 60, 80, 90], "text": "1001", "confidence": 0.95, "line_no": 3, "page_no": 1, "source_path": "demo.png"},
                    {"block_id": "b4", "block_type": "text", "bbox": [120, 60, 230, 90], "text": "12.5", "confidence": 0.95, "line_no": 4, "page_no": 1, "source_path": "demo.png"},
                ],
                "paddleocr",
            )
            rows, meta = extract_image_rows("demo.png", spec={})

        self.assertEqual(len(rows), 4)
        self.assertGreaterEqual(len(meta["table_cells"]), 4)
        self.assertEqual(meta["table_cells"][0]["text"], "编号")

    def test_extract_xlsx_rows_prefers_docling_when_available(self):
        with patch("aiwf.ingest_xlsx_pipeline.extract_with_docling") as extract_docling:
            extract_docling.return_value = {
                "rows": [
                    {
                        "source_file": "a.xlsx",
                        "source_path": "a.xlsx",
                        "source_type": "xlsx",
                        "sheet_name": "Sheet1",
                        "row_index": 1,
                        "amount": 10.0,
                    }
                ],
                "sheet_frames": [
                    {
                        "workbook_name": "a.xlsx",
                        "sheet_name": "Sheet1",
                        "sheet_index": 0,
                        "header_row_span": [1, 1],
                        "header_confidence": 1.0,
                        "table_name": "",
                        "columns": ["amount"],
                        "row_count": 1,
                        "blank_rows": 0,
                        "numeric_cells_total": 1,
                        "numeric_cells_parsed": 1,
                        "date_cells_total": 0,
                        "date_cells_parsed": 0,
                        "formula_cells": 0,
                        "formula_mismatches": 0,
                        "hidden": False,
                        "source_path": "a.xlsx",
                    }
                ],
                "table_cells": [{"cell_id": "c1", "row": 1, "col": 1, "text": "10"}],
                "engine_trace": [{"engine": "docling", "ok": True}],
            }
            rows, meta = extract_xlsx_rows("a.xlsx", spec={"xlsx_rules": {"required_columns": ["amount"]}})

        self.assertEqual(len(rows), 1)
        self.assertEqual(meta["engine"], "docling")
        self.assertEqual(len(meta["table_cells"]), 1)
        self.assertFalse(meta["quality_blocked"])

    def test_load_pdf_input_prefers_docling_metadata_when_available(self):
        with patch("aiwf.ingest_file_readers.extract_with_docling") as extract_docling:
            extract_docling.return_value = {
                "image_blocks": [
                    {
                        "block_id": "doc_blk_0001",
                        "block_type": "text",
                        "text": "Claim: policy reform improves compliance and reduces friction.",
                        "page_no": 1,
                        "line_no": 1,
                        "bbox": [0, 0, 10, 10],
                    }
                ],
                "table_cells": [
                    {"cell_id": "doc_cell_1", "row": 1, "col": 1, "text": "Acct No"},
                    {"cell_id": "doc_cell_2", "row": 1, "col": 2, "text": "Posting Dt"},
                    {"cell_id": "doc_cell_3", "row": 2, "col": 1, "text": "62220001"},
                    {"cell_id": "doc_cell_4", "row": 2, "col": 2, "text": "2026/03/01"},
                ],
                "sheet_frames": [],
                "engine_trace": [{"engine": "docling", "ok": True}],
            }
            rows, meta = load_pdf_input("demo.pdf", {"text_by_line": False})

        self.assertEqual(meta["input_format"], "pdf")
        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["source_type"], "pdf")
        self.assertEqual(len(meta["table_cells"]), 4)
        self.assertEqual(meta["engine_trace"][0]["engine"], "docling")

    def test_extract_xlsx_rows_handles_chinese_multirow_headers_and_units(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "cn_finance.xlsx")
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "财务"
            ws["A1"] = "编号"
            ws.merge_cells("A1:A2")
            ws["B1"] = "收款信息"
            ws.merge_cells("B1:C1")
            ws["B2"] = "金额（万元）"
            ws["C2"] = "业务日期"
            ws.append([None, None, None])
            ws.append([1001, "12.5", "2026年3月1日"])
            ws.append([1002, "2", "2026/03/02"])
            wb.save(p)

            rows, meta = extract_xlsx_rows(
                p,
                spec={
                    "canonical_profile": "finance_statement",
                    "xlsx_rules": {
                        "required_columns": ["id", "amount", "biz_date"],
                        "header_confidence_min": 0.8,
                    },
                },
            )

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0]["id"], 1001)
        self.assertEqual(rows[0]["amount"], 125000.0)
        self.assertEqual(rows[0]["biz_date"], "2026-03-01")
        self.assertFalse(meta["quality_blocked"])

    def test_extract_xlsx_rows_handles_bank_statement_multiheaders(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "bank_statement.xlsx")
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "流水"
            ws["A1"] = "账号"
            ws.merge_cells("A1:A2")
            ws["B1"] = "交易日期"
            ws.merge_cells("B1:B2")
            ws["C1"] = "借贷信息"
            ws.merge_cells("C1:D1")
            ws["C2"] = "借方金额"
            ws["D2"] = "贷方金额"
            ws["E1"] = "金额"
            ws.merge_cells("E1:E2")
            ws["F1"] = "余额（万元）"
            ws.merge_cells("F1:F2")
            ws["G1"] = "对方户名"
            ws.merge_cells("G1:G2")
            ws["H1"] = "流水号"
            ws.merge_cells("H1:H2")
            ws.append(["62220001", "2026年3月1日", "120.5", "0", "-120.5", "1.25", "张三", "TXN-001"])
            wb.save(p)

            rows, meta = extract_xlsx_rows(
                p,
                spec={
                    "canonical_profile": "bank_statement",
                    "quality_rules": {
                        "canonical_profile": "bank_statement",
                        "required_fields": ["account_no", "txn_date"],
                        "unique_keys": ["account_no", "txn_date", "ref_no", "amount"],
                        "max_required_missing_ratio": 0.0,
                        "numeric_parse_rate_min": 1.0,
                        "date_parse_rate_min": 1.0,
                    },
                    "xlsx_rules": {
                        "required_columns": ["account_no", "txn_date", "debit_amount", "credit_amount", "amount", "balance", "counterparty_name", "ref_no"],
                        "header_confidence_min": 0.8,
                    },
                },
            )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0]["account_no"], "62220001")
        self.assertEqual(rows[0]["txn_date"], "2026-03-01")
        self.assertEqual(rows[0]["debit_amount"], 120.5)
        self.assertEqual(rows[0]["amount"], -120.5)
        self.assertEqual(rows[0]["balance"], 12500.0)
        self.assertEqual(rows[0]["counterparty_name"], "张三")
        self.assertEqual(rows[0]["ref_no"], "TXN-001")
        self.assertFalse(meta["quality_blocked"])

    def test_extract_xlsx_rows_auto_mode_handles_bank_statement_abbrev_headers(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "bank_statement_abbrev.xlsx")
            from openpyxl import Workbook  # type: ignore

            wb = Workbook()
            ws = wb.active
            ws.title = "BankFlow"
            ws.append(["Acct No", "Posting Dt", "DR", "CR", "Bal", "Memo", "Ref No"])
            ws.append(["62220001", "2026/03/01", "120.5", "0", "12500", "Rent", "TXN-001"])
            ws.append(["62220001", "2026-03-02", "0", "300", "12800", "Refund", "TXN-002"])
            wb.save(p)

            strict_rows, strict_meta = extract_xlsx_rows(
                p,
                spec={
                    "canonical_profile": "bank_statement",
                    "header_mapping_mode": "strict",
                    "xlsx_rules": {
                        "required_columns": ["account_no", "txn_date", "debit_amount", "credit_amount", "balance", "remark", "ref_no"],
                        "header_confidence_min": 0.8,
                    },
                },
            )
            auto_rows, auto_meta = extract_xlsx_rows(
                p,
                spec={
                    "canonical_profile": "bank_statement",
                    "header_mapping_mode": "auto",
                    "xlsx_rules": {
                        "required_columns": ["account_no", "txn_date", "debit_amount", "credit_amount", "balance", "remark", "ref_no"],
                        "header_confidence_min": 0.8,
                    },
                },
            )

        self.assertEqual(len(strict_rows), 2)
        self.assertTrue(strict_meta["quality_blocked"])
        self.assertEqual(len(auto_rows), 2)
        self.assertEqual(auto_rows[0]["account_no"], "62220001")
        self.assertEqual(auto_rows[0]["txn_date"], "2026-03-01")
        self.assertEqual(auto_rows[0]["debit_amount"], 120.5)
        self.assertEqual(auto_rows[0]["credit_amount"], 0.0)
        self.assertEqual(auto_rows[0]["balance"], 12500.0)
        self.assertEqual(auto_rows[0]["remark"], "Rent")
        self.assertEqual(auto_rows[0]["ref_no"], "TXN-001")
        self.assertFalse(auto_meta["quality_blocked"])

    def test_read_docx(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.docx")
            from docx import Document  # type: ignore

            d = Document()
            d.add_paragraph("debate evidence paragraph")
            d.save(p)

            rows, meta = ingest.load_rows_from_file(p)
            self.assertEqual(meta["input_format"], "docx")
            self.assertGreaterEqual(len(rows), 1)
            self.assertIn("debate", rows[0]["text"])

    def test_skip_image_when_ocr_disabled(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.png")
            with open(p, "wb") as f:
                f.write(b"not-a-real-image")
            rows, meta = ingest.load_rows_from_file(p, ocr_enabled=False)
            self.assertEqual(rows, [])
            self.assertEqual(meta["input_format"], "image")
            self.assertTrue(meta.get("skipped"))

    def test_real_ocr_image_assets_exist_and_are_readable(self):
        from PIL import Image  # type: ignore

        root = Path(__file__).resolve().parents[3] / "lake" / "datasets" / "regression_v1_2_sidecar_gold"
        paths = [
            root / "image_bank_statement_tabular_auto" / "input.png",
            root / "image_debate_text_auto" / "input.png",
            root / "image_customer_contact_auto" / "input.png",
            root / "image_customer_ledger_auto" / "input.png",
        ]
        for path in paths:
            self.assertTrue(path.exists(), str(path))
            with Image.open(path) as image:
                self.assertGreater(image.width, 100)
                self.assertGreater(image.height, 100)

    def test_real_ocr_pdf_assets_exist_and_are_readable(self):
        from pypdf import PdfReader  # type: ignore

        root = Path(__file__).resolve().parents[3] / "lake" / "datasets" / "regression_v1_2_sidecar_gold"
        paths = [
            root / "pdf_bank_statement_tabular_auto" / "input.pdf",
            root / "pdf_debate_text_auto" / "input.pdf",
            root / "pdf_customer_contact_auto" / "input.pdf",
            root / "pdf_customer_ledger_auto" / "input.pdf",
        ]
        for path in paths:
            self.assertTrue(path.exists(), str(path))
            reader = PdfReader(str(path))
            self.assertGreaterEqual(len(reader.pages), 1)

    def test_register_input_reader_supports_custom_extension(self):
        with tempfile.TemporaryDirectory() as tmp:
            p = os.path.join(tmp, "a.custom")
            with open(p, "w", encoding="utf-8") as f:
                f.write("ignored")

            def load_custom(path, options):
                return (
                    [{"text": "custom", "source_path": path, "source_type": "custom"}],
                    {"input_format": "custom"},
                )

            ingest.register_input_reader(
                "custom",
                [".custom"],
                load_custom,
                domain="custom-ingest",
                domain_metadata={"label": "Custom Ingest", "backend": "extension", "builtin": False},
            )
            try:
                rows, meta = ingest.load_rows_from_file(p)
                details = {item["input_format"]: item for item in ingest.list_input_reader_details()}
                domains = ingest.list_input_reader_domains()
            finally:
                ingest.unregister_input_reader("custom")

            self.assertEqual(meta["input_format"], "custom")
            self.assertEqual(rows[0]["source_type"], "custom")
            self.assertEqual(details["custom"]["domain"], "custom-ingest")
            self.assertTrue(any(item["name"] == "custom-ingest" for item in domains))

    def test_register_input_reader_rejects_conflicting_extension_when_requested(self):
        def load_markdown(path, options):
            return [{"text": "markdown"}], {"input_format": "markdown"}

        with self.assertRaises(RuntimeError):
            ingest.register_input_reader("markdown", [".txt"], load_markdown, on_conflict="error")

    def test_register_input_reader_can_replace_extension_owner(self):
        def load_markdown(path, options):
            return [{"text": "markdown"}], {"input_format": "markdown"}

        original = ingest.get_input_reader("demo.txt")
        ingest.register_input_reader("markdown", [".txt"], load_markdown, on_conflict="warn")
        try:
            self.assertEqual(ingest.get_input_reader("demo.txt").input_format, "markdown")
            details = {item["input_format"]: item for item in ingest.list_input_reader_details()}
            self.assertEqual(details["markdown"]["extensions"], [".txt"])
            self.assertNotIn("txt", details)
        finally:
            ingest.unregister_input_reader("markdown")
            ingest.register_input_reader(
                original.input_format,
                original.extensions,
                original.loader,
                source_module=original.source_module,
                on_conflict="replace",
            )


if __name__ == "__main__":
    unittest.main()
